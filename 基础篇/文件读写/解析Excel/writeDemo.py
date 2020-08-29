from openpyxl import Workbook
from openpyxl.utils import get_column_letter

rows = [{'title': 'unex-pim.itemOrderRule.addItemOrderRuleByList', 'titleName': 'PLP排序自定义得分接口', 'server': '商品服务',
         'req': '{\n    "condition": "[{\\"spuCode\\":\\"001\\",\\"spuScore\\":0.2,\\"ruleCode\\":\\"RULE_A\\"}]",\n    "tenantCode": "880001"\n}',
         'res': '{\n    "sign": "5bf49fd7f7f139750dc94fb35a244016",\n    "unex-pim_itemOrderRule_addItemOrderRuleByList_response": {\n        "msg": "授权权限不足",\n        "code": "2001",\n        "subCode": "isv.api-access-denied",\n        "subMsg": "api 访问权限不足"\n    }\n}'},
        {'title': 'unex-pim.itemOrderRule.deleteItemOrderRule', 'titleName': 'PLP排序规则删除接口', 'server': '商品服务',
         'req': '{\n    "ruleCode": "RULE_C",\n    "ruleId": "1",\n    "tenantCode": "880001"\n}',
         'res': '{\n    "unex-pim_itemOrderRule_deleteItemOrderRule_response": {\n        "msg": "授权权限不足",\n        "code": "2001",\n        "subCode": "isv.api-access-denied",\n        "subMsg": "api 访问权限不足"\n    },\n    "sign": "5bf49fd7f7f139750dc94fb35a244016"\n}'}]

wb = Workbook()
ws1 = wb.active
# ws1.title = "range names"
# for row in range(1, 40):
#     ws1.append(range(600))

# add new sheet
ws2 = wb.create_sheet(title="Pi")
ws2['A1'] = '请求名称'
ws2['B1'] = '接口描述'
ws2['C1'] = '服务名称'
ws2['D1'] = '请求'
ws2['E1'] = '响应'

for index, row in enumerate(rows):
    start = 1
    for k, v in dict(row).items():
        print(k, v)
        ws2.cell(column=start, row=index + 2, value=v)
        start += 1


# add new sheet
# ws3 = wb.create_sheet(title="Data")
#
# for row in range(10, 20):
#     for col in range(27, 54):
#         _ = ws3.cell(column=col, row=row, value="{0}".format(get_column_letter(col)))
# print(ws3['AA10'].value)
#
# # save excel
dest_filename = 'empty_book.xlsx'
wb.save(filename=dest_filename)
